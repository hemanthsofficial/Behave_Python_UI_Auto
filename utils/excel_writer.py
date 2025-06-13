import os

import openpyxl

from utils.logger import get_logger

logger = get_logger(__name__)


class ExcelWriter:
    @staticmethod
    def write_to_excel(log_file_path, sheet_name, data=None):
        if data is None:
            data = []
        try:
            excel_dir = os.path.join(os.getcwd(), "excel")
            os.makedirs(excel_dir, exist_ok=True)

            filename_only = os.path.basename(log_file_path).replace(".log", ".xlsx")
            excel_path = os.path.join(excel_dir, filename_only)

            logger.info(f"Writing to Excel: {excel_path}, Sheet: {sheet_name}")
            if os.path.exists(excel_path):
                workbook = openpyxl.load_workbook(excel_path)
            else:
                workbook = openpyxl.Workbook()
                workbook.remove(workbook.active)

            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]

            sheet = workbook.create_sheet(title=sheet_name)
            for row in data:
                sheet.append(row)

            workbook.save(excel_path)
            logger.info(f"Excel saved successfully: {excel_path}")
        except Exception as e:
            logger.error(f"Failed to write to Excel file: {e}", exc_info=True)
            raise
